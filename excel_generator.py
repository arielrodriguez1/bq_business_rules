"""
excel_generator.py
------------------
Módulo de generación del archivo Excel con reglas de negocio.
Aplica estilos, anchos de columna y formato condicional.
"""

import os
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
#  Ruta del template base
# ──────────────────────────────────────────────
TEMPLATE_PATH = Path(r"C:\Users\vn59q13\OneDrive - Walmart Inc\Documentos\bq_business_rules\Reglas_Negocio_template.xlsx")


# ──────────────────────────────────────────────
#  Paleta de colores Walmart
# ──────────────────────────────────────────────
WALMART_BLUE   = "0071CE"   # Azul corporativo
WALMART_YELLOW = "FFC220"   # Amarillo corporativo
HEADER_BG      = WALMART_BLUE
HEADER_FONT    = "FFFFFF"
ROW_ALT        = "EEF4FB"   # Azul muy claro para filas alternas
NULL_YES_COLOR = "D9EAD3"   # Verde suave  → Nulo permitido
NULL_NO_COLOR  = "FCE5CD"   # Naranja suave → Nulo NO permitido
TO_FILL_COLOR  = "FFF9E6"   # Amarillo muy suave → Campos por completar

# Anchos minimos por columna (caracteres)
COL_WIDTHS = {
    "Columna":                       25,
    "Descripcion":                   50,
    "Tabla Origen":                  45,
    "Tipo":                          18,
    "Formato":                       45,
    "Valores Permitidos":            55,
    "Valor Nulo Permitido":          22,
    "Criticidad":                    20,
    "Justificacion de Criticidad":   50,
    "Ejemplo":                       35,
    "Clasificacion de Sensibilidad":  30,
    "Referencia Standard":           60,
    "Observacion":                   80,
}


def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style() -> dict:
    return {
        "font":      Font(name="Calibri", bold=True, color=HEADER_FONT, size=11),
        "fill":      PatternFill("solid", fgColor=HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    _thin_border(),
    }


def _apply_style(cell, style: dict):
    for attr, value in style.items():
        setattr(cell, attr, value)


# ──────────────────────────────────────────────
#  Función principal
# ──────────────────────────────────────────────
def generate_excel(
    df: pd.DataFrame,
    project_id: str,
    dataset_id: str,
    table_id: str,
    output_dir: str,
) -> str:
    """
    Recibe el DataFrame de reglas de negocio y genera el archivo Excel.
    Utiliza el template base para mantener formato consistente.

    Retorna la ruta absoluta del archivo generado.
    """
    os.makedirs(output_dir, exist_ok=True)

    # Verificar que el template exista
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template no encontrado: {TEMPLATE_PATH}\n"
            f"Por favor verifica que el archivo exista."
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"Reglas_Negocio_{table_id}_{timestamp}.xlsx"
    filepath  = os.path.join(output_dir, filename)

    # ── 1. Copiar el template como base ──────────────────────────────────────
    print(f"[i] Usando template: {TEMPLATE_PATH.name}")
    shutil.copy2(TEMPLATE_PATH, filepath)

    # ── 2. Cargar el archivo copiado ─────────────────────────────────────────
    wb = load_workbook(filepath)
    
    # Verificar que las hojas existan
    sheet_name = "Reglas de Negocio"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el template")
    
    ws = wb[sheet_name]

    # ── 3. Limpiar datos previos (mantener solo header) ──────────────────
    # Eliminar todas las filas de datos, pero mantener la primera (cabecera)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    
    # ── 4. Escribir los datos ─────────────────────────────────────────────
    columns = list(df.columns)
    num_cols = len(columns)
    
    # Escribir datos fila por fila (empezando desde la fila 2)
    for df_row_idx, row_data in df.iterrows():
        excel_row = df_row_idx + 2  # +2 porque excel es 1-indexed y fila 1 es header
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = row_data[col_name]
    
    num_rows = len(df) + 1   # +1 por la cabecera

    # ── 5. Aplicar estilos a las filas de datos ──────────────────────────
    null_col_idx = columns.index("Valor Nulo Permitido") + 1
    criticidad_col_idx = columns.index("Criticidad") + 1 if "Criticidad" in columns else -1
    sensibilidad_col_idx = columns.index("Clasificacion de Sensibilidad") + 1 if "Clasificacion de Sensibilidad" in columns else -1

    for row_idx in range(2, num_rows + 1):
        is_alt = (row_idx % 2 == 0)
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Fondo alterno
            bg = ROW_ALT if is_alt else "FFFFFF"
            
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Calibri", size=10)
            cell.border    = _thin_border()
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

        # Color especial para "Valor Nulo Permitido"
        null_cell = ws.cell(row=row_idx, column=null_col_idx)
        null_val  = str(null_cell.value).strip() if null_cell.value else ""
        if null_val == "Si":
            null_cell.fill = PatternFill("solid", fgColor=NULL_YES_COLOR)
            null_cell.font = Font(name="Calibri", size=10, color="274E13")
        elif null_val == "No":
            null_cell.fill = PatternFill("solid", fgColor=NULL_NO_COLOR)
            null_cell.font = Font(name="Calibri", size=10, color="7F2B00")
        null_cell.alignment = Alignment(horizontal="center", vertical="top")
        
        # Color especial para "Criticidad"
        if criticidad_col_idx > 0:
            crit_cell = ws.cell(row=row_idx, column=criticidad_col_idx)
            crit_val = str(crit_cell.value).strip() if crit_cell.value else ""
            if crit_val == "Alta":
                crit_cell.fill = PatternFill("solid", fgColor="F4CCCC")  # Rojo suave
                crit_cell.font = Font(name="Calibri", size=10, bold=True, color="990000")
            elif crit_val == "Media":
                crit_cell.fill = PatternFill("solid", fgColor="FCE5CD")  # Naranja suave
                crit_cell.font = Font(name="Calibri", size=10, bold=True, color="E69138")
            elif crit_val == "Baja":
                crit_cell.fill = PatternFill("solid", fgColor="D9EAD3")  # Verde suave
                crit_cell.font = Font(name="Calibri", size=10, color="274E13")
            crit_cell.alignment = Alignment(horizontal="center", vertical="top")
        
        # Color especial para "Clasificacion de Sensibilidad"
        if sensibilidad_col_idx > 0:
            sens_cell = ws.cell(row=row_idx, column=sensibilidad_col_idx)
            sens_val = str(sens_cell.value).strip() if sens_cell.value else ""
            if sens_val == "Highly Sensitive":
                sens_cell.fill = PatternFill("solid", fgColor="EA9999")  # Rojo más intenso
                sens_cell.font = Font(name="Calibri", size=10, bold=True, color="990000")
            elif sens_val == "Sensitive":
                sens_cell.fill = PatternFill("solid", fgColor="F9CB9C")  # Naranja más intenso
                sens_cell.font = Font(name="Calibri", size=10, bold=True, color="E69138")
            elif sens_val == "Non Sensitive":
                sens_cell.fill = PatternFill("solid", fgColor="B6D7A8")  # Verde más intenso
                sens_cell.font = Font(name="Calibri", size=10, color="274E13")
            sens_cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[row_idx].height = 50

    # ── 6. Actualizar pestaña de metadatos ──────────────────────────────────────
    # Eliminar y recrear la hoja para evitar problemas con celdas fusionadas
    if "Metadatos" in wb.sheetnames:
        wb.remove(wb["Metadatos"])
    ws_meta = wb.create_sheet("Metadatos")
    meta_info = [
        ("Proyecto GCP",  project_id),
        ("Dataset",       dataset_id),
        ("Tabla",         table_id),
        ("Total columnas", len(df)),
        ("Fecha generación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Generado por",  "bq_business_rules - Wibey/Walmart"),
    ]
    for r_idx, (key, val) in enumerate(meta_info, start=1):
        key_cell = ws_meta.cell(row=r_idx, column=1, value=key)
        key_cell.font      = Font(name="Calibri", bold=True, size=11, color=WALMART_BLUE)
        key_cell.alignment = Alignment(horizontal="right")
        val_cell = ws_meta.cell(row=r_idx, column=2, value=val)
        val_cell.font      = Font(name="Calibri", size=11)

    ws_meta.column_dimensions["A"].width = 22
    ws_meta.column_dimensions["B"].width = 50
    
    # -- 7. Actualizar Resumen por Clasificacion de Sensibilidad --------------
    # Eliminar y recrear la hoja para evitar problemas con celdas fusionadas
    if "Resumen por Clasificacion" in wb.sheetnames:
        wb.remove(wb["Resumen por Clasificacion"])
    ws_summary = wb.create_sheet("Resumen por Clasificacion", 1)  # Insert as 2nd sheet
    
    # Agrupar columnas por clasificación
    grouped = {}
    for _, row in df.iterrows():
        classification = row.get("Clasificacion de Sensibilidad", "Non Sensitive")
        if classification not in grouped:
            grouped[classification] = []
        grouped[classification].append(row)
    
    # Orden de presentación (de más sensible a menos)
    classification_order = ["Highly Sensitive", "Sensitive", "Non Sensitive"]
    
    current_row = 1
    
    # Título principal
    title_cell = ws_summary.cell(row=current_row, column=1)
    title_cell.value = "RESUMEN DE CLASIFICACION DE SENSIBILIDAD DE DATOS"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=WALMART_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.merge_cells(f"A{current_row}:E{current_row}")
    current_row += 1
    
    subtitle_cell = ws_summary.cell(row=current_row, column=1)
    subtitle_cell.value = f"Standard: DC-DG-03-02 - Global Data Sensitivity Classification | Tabla: {table_id}"
    subtitle_cell.font = Font(name="Calibri", italic=True, size=10, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="center")
    ws_summary.merge_cells(f"A{current_row}:E{current_row}")
    current_row += 2
    
    # Estadísticas generales
    stats_start = current_row
    stats_data = [
        ("Total de Columnas:", len(df)),
        ("Highly Sensitive:", len(grouped.get("Highly Sensitive", []))),
        ("Sensitive:", len(grouped.get("Sensitive", []))),
        ("Non Sensitive:", len(grouped.get("Non Sensitive", []))),
    ]
    
    for label, value in stats_data:
        label_cell = ws_summary.cell(row=current_row, column=1, value=label)
        label_cell.font = Font(name="Calibri", bold=True, size=11)
        label_cell.alignment = Alignment(horizontal="right")
        
        value_cell = ws_summary.cell(row=current_row, column=2, value=value)
        value_cell.font = Font(name="Calibri", size=11, bold=True)
        value_cell.alignment = Alignment(horizontal="left")
        
        # Color según clasificación
        if "Highly" in label:
            value_cell.font = Font(name="Calibri", size=11, bold=True, color="990000")
        elif "Sensitive:" in label and "Non" not in label:
            value_cell.font = Font(name="Calibri", size=11, bold=True, color="E69138")
        elif "Non Sensitive" in label:
            value_cell.font = Font(name="Calibri", size=11, bold=True, color="274E13")
        
        current_row += 1
    
    current_row += 2
    
    # Tabla detallada por clasificación
    for classification in classification_order:
        if classification not in grouped or len(grouped[classification]) == 0:
            continue
        
        # Título de sección
        section_cell = ws_summary.cell(row=current_row, column=1)
        section_cell.value = f"{classification.upper()} ({len(grouped[classification])} columnas)"
        section_cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color de fondo según clasificación
        if classification == "Highly Sensitive":
            bg_color = "990000"  # Rojo oscuro
        elif classification == "Sensitive":
            bg_color = "E69138"  # Naranja
        else:
            bg_color = "274E13"  # Verde oscuro
        
        section_cell.fill = PatternFill("solid", fgColor=bg_color)
        ws_summary.merge_cells(f"A{current_row}:E{current_row}")
        ws_summary.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Headers de la tabla
        headers = ["Columna", "Tipo", "Referencia Standard", "Observacion", "Ejemplo"]
        for col_idx, header in enumerate(headers, start=1):
            header_cell = ws_summary.cell(row=current_row, column=col_idx, value=header)
            header_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            header_cell.fill = PatternFill("solid", fgColor=WALMART_BLUE)
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_cell.border = _thin_border()
        ws_summary.row_dimensions[current_row].height = 30
        current_row += 1
        
        # Filas de datos
        for row_data in grouped[classification]:
            row_values = [
                row_data.get("Columna", ""),
                row_data.get("Tipo", ""),
                row_data.get("Referencia Standard", ""),
                row_data.get("Observacion", ""),
                row_data.get("Ejemplo", "")[:50] if row_data.get("Ejemplo", "") else "",  # Truncar ejemplo
            ]
            
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws_summary.cell(row=current_row, column=col_idx, value=value)
                cell.font = Font(name="Calibri", size=9)
                cell.border = _thin_border()
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Fondo alterno
                if current_row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F9F9F9")
            
            ws_summary.row_dimensions[current_row].height = 40
            current_row += 1
        
        current_row += 2  # Espacio entre secciones
    
    # Ajustar anchos de columnas en resumen
    ws_summary.column_dimensions["A"].width = 25  # Columna
    ws_summary.column_dimensions["B"].width = 15  # Tipo
    ws_summary.column_dimensions["C"].width = 60  # Referencia Standard
    ws_summary.column_dimensions["D"].width = 85  # Observacion
    ws_summary.column_dimensions["E"].width = 35  # Ejemplo

    # ── 8. Guardar el archivo ────────────────────────────────────────────
    wb.save(filepath)
    
    print(f"\n[OK] Excel generado exitosamente:")
    print(f"    {filepath}\n")
    return filepath
